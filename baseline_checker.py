import xlwings as xw
import openpyxl as xl



class Output:
    def __init__(self,curr_monit, avanterra):
        self.avantera = avanterra
        self.curr_monit = curr_monit

    def col_avanterra(self):
        wb_av = xl.load_workbook(self.avantera)
        sheet_av = wb_av.active
        head_av = {cell.value: cell.column for cell in sheet_av[1]} #pobranie naglowkow
        columns_name_av = ['Isin','A']
        data_av = []
        for row in sheet_av.iter_rows(min_row=2, values_only=True):
            data_av.append({name: row[head_av[name]-1] for name in columns_name_av}) #pobranei slownika z danych po nazwie kolumn

    def curr_monit(self):
        wb_curr_monit = xl.load_workbook(self.curr_monit)
        sheet_cm = wb_curr_monit.active
        head_cm = {cell.value: cell.column for cell in sheet_cm[1]}
        columns_name_cm = ['Isin', 'B']
        data_cm = []
        for row in sheet_cm.iter_rows(min_row=2, values_only=True):
            data_cm.append({name: row[head_cm[name]-1] for name in columns_name_cm})

    def new_book:
        wb = xw.Book()
        sheet = wb.sheet[0]
        headers_cm = list(d)

class Sources:
    def __init__(self, avanterra, curr_monit):
        self.avanterra = C:\Users\Paulina\PycharmProjects\Exercises\baseline\input
        self.curr_monit = C:\Users\Paulina\PycharmProjects\Exercises\baseline\input
